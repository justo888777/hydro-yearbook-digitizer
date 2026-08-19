from __future__ import annotations

import re
import calendar
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

_NUMERIC = re.compile(r"^-?\d+(?:\.\d+)?$")


@dataclass(frozen=True)
class OcrToken:
    text: str
    x: float
    y: float
    confidence: float


@dataclass(frozen=True)
class DailyCandidate:
    source_file: str
    table_id: str
    day: int
    month: int
    raw_text: str
    value: float
    x: float
    y: float
    confidence: float
    status: str = "needs_review"
    engine: str = "rapidocr_onnxruntime"


def constant_daily_candidates(
    *, source_file: str, table_id: str, year: int, value: float, engine: str
) -> list[DailyCandidate]:
    """Create a review-only transcription when two readings establish one constant printed matrix."""
    return [
        DailyCandidate(source_file, table_id, day, month, str(value), value, 0, 0, 1.0, engine=engine)
        for month in range(1, 13)
        for day in range(1, calendar.monthrange(year, month)[1] + 1)
    ]


def normalized_crop(image_path: Path, output_path: Path, region: tuple[float, float, float, float]) -> Path:
    """Write a derived crop from immutable input using normalized coordinates."""
    try:
        from PIL import Image, ImageOps
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Pillow is required") from exc
    left, top, right, bottom = region
    if not (0 <= left < right <= 1 and 0 <= top < bottom <= 1):
        raise ValueError("crop region must be four normalized coordinates between 0 and 1")
    with Image.open(image_path) as source:
        image = ImageOps.exif_transpose(source)
        crop = image.crop((round(left * image.width), round(top * image.height), round(right * image.width), round(bottom * image.height)))
        output_path.parent.mkdir(parents=True, exist_ok=True)
        crop.save(output_path, quality=95)
    return output_path


def normalized_quad_crop(
    image_path: Path,
    output_path: Path,
    corners: tuple[float, float, float, float, float, float, float, float],
    *,
    target_size: tuple[int, int] = (2800, 1600),
) -> Path:
    """Perspective-rectify a logical table using normalized corner coordinates."""
    try:
        import cv2
        import numpy as np
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("OpenCV and numpy are required") from exc
    encoded = np.fromfile(image_path, dtype=np.uint8)
    image = cv2.imdecode(encoded, cv2.IMREAD_COLOR)
    if image is None:
        raise ValueError(f"cannot decode image: {image_path}")
    points = np.array(corners, dtype=np.float32).reshape(4, 2)
    if np.any(points < 0) or np.any(points > 1):
        raise ValueError("quad coordinates must be normalized between 0 and 1")
    points[:, 0] *= image.shape[1]
    points[:, 1] *= image.shape[0]
    width, height = target_size
    target = np.array([[0, 0], [width - 1, 0], [width - 1, height - 1], [0, height - 1]], dtype=np.float32)
    rectified = cv2.warpPerspective(image, cv2.getPerspectiveTransform(points, target), target_size)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    cv2.imencode(".jpg", rectified)[1].tofile(output_path)
    return output_path


def read_ocr_tokens(image_path: Path) -> list[OcrToken]:
    """Read raw local OCR candidates without normalizing or accepting any value."""
    try:
        from rapidocr_onnxruntime import RapidOCR
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("RapidOCR is required: pip install -e .[ocr]") from exc
    result, _ = RapidOCR()(str(image_path))
    tokens = []
    for box, text, confidence in result or []:
        x = sum(point[0] for point in box) / len(box)
        y = sum(point[1] for point in box) / len(box)
        tokens.append(OcrToken(str(text).strip(), x, y, float(confidence)))
    return tokens


def map_daily_tokens(
    tokens: list[OcrToken],
    *,
    source_file: str,
    table_id: str,
    month_centers: tuple[float, ...],
    day_column_max_x: float,
    daily_y_min: float,
    daily_y_max: float,
    y_tolerance: float = 14,
    x_tolerance: float = 78,
    row_start: float | None = None,
    row_step: float | None = None,
    year: int | None = None,
) -> tuple[list[DailyCandidate], list[OcrToken]]:
    """Map OCR tokens onto a daily matrix; ambiguous cells remain review items."""
    day_tokens = [
        token for token in tokens
        if token.x <= day_column_max_x and daily_y_min <= token.y <= daily_y_max and token.text.isdigit() and 1 <= int(token.text) <= 31
    ]
    day_centers: dict[int, float] = {}
    for token in day_tokens:
        day = int(token.text)
        if day not in day_centers or abs(token.y - day_centers[day]) < 1:
            day_centers[day] = token.y
    if row_start is not None and row_step is not None:
        day_centers = {day: row_start + (day - 1) * row_step for day in range(1, 32)}
    elif len(day_centers) < 28:
        return [], tokens

    mapped: dict[tuple[int, int], list[OcrToken]] = {}
    leftovers: list[OcrToken] = []
    for token in tokens:
        if token.x <= day_column_max_x:
            continue
        if not _NUMERIC.fullmatch(token.text) or not (daily_y_min <= token.y <= daily_y_max):
            leftovers.append(token)
            continue
        month_index = min(range(12), key=lambda index: abs(token.x - month_centers[index]))
        day = min(day_centers, key=lambda candidate: abs(token.y - day_centers[candidate]))
        if abs(token.x - month_centers[month_index]) > x_tolerance or abs(token.y - day_centers[day]) > y_tolerance:
            leftovers.append(token)
            continue
        mapped.setdefault((day, month_index + 1), []).append(token)

    records: list[DailyCandidate] = []
    for (day, month), candidates in sorted(mapped.items()):
        if year is not None and day > calendar.monthrange(year, month)[1]:
            leftovers.extend(candidates)
            continue
        if len(candidates) != 1:
            leftovers.extend(candidates)
            continue
        token = candidates[0]
        records.append(DailyCandidate(source_file, table_id, day, month, token.text, float(token.text), token.x, token.y, token.confidence))
    return records, leftovers


def write_daily_candidate_workbook(
    output_path: Path,
    *,
    basin: str,
    year: int,
    table_id: str,
    source_file: str,
    crop_path: Path,
    records: list[DailyCandidate],
    leftovers: list[OcrToken],
) -> None:
    """Export a clearly labelled candidate workbook; no record is auto-released."""
    book = Workbook()
    metadata = book.active
    metadata.title = "元数据"
    metadata.append(["字段", "值"])
    metadata.append(["流域", basin])
    metadata.append(["年份", year])
    metadata.append(["表格ID", table_id])
    metadata.append(["来源图片", source_file])
    metadata.append(["派生表格图", str(crop_path)])
    metadata.append(["识别引擎", "rapidocr_onnxruntime"])
    metadata.append(["发布状态", "blocked_pending_second_engine_and_review"])

    daily = book.create_sheet("逐日数据")
    daily.append(["日"] + list(range(1, 13)))
    lookup = {(record.day, record.month): record.value for record in records}
    for day in range(1, 32):
        daily.append([day] + [lookup.get((day, month)) for month in range(1, 13)])

    long = book.create_sheet("长表")
    long.append(["month", "day", "value", "raw_text", "status", "source_file", "table_id", "x", "y", "confidence", "engine"])
    for record in records:
        long.append([record.month, record.day, record.value, record.raw_text, record.status, record.source_file, record.table_id, record.x, record.y, record.confidence, record.engine])

    monthly = book.create_sheet("月统计")
    monthly.append(["month", "numeric_count", "mean", "maximum", "minimum", "status"])
    for month in range(1, 13):
        values = [record.value for record in records if record.month == month]
        monthly.append([month, len(values), sum(values) / len(values) if values else None, max(values) if values else None, min(values) if values else None, "needs_review"])

    audit = book.create_sheet("识别审计")
    audit.append(["raw_text", "x", "y", "confidence", "reason"])
    for token in leftovers:
        audit.append([token.text, token.x, token.y, token.confidence, "unmapped_or_ambiguous"])

    review = book.create_sheet("待审核")
    review.append(["item", "status", "reason"])
    review.append(["all exported daily values", "blocking", "single local OCR engine; second engine and visual review required"])
    review.append(["unmapped OCR tokens", "blocking" if leftovers else "none", f"{len(leftovers)} tokens require review"])

    for sheet in book.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(46, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    book.save(output_path)
