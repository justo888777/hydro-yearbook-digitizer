from __future__ import annotations

import csv
import json
import math
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

from .page_classes import PageClass, route_page
from .project import sha256_file
from .project import ProjectValidation


_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp", ".gif"}


@dataclass(frozen=True)
class SourceRecord:
    source_id: str
    source_file: str
    source_path: str
    sha256: str
    width: int
    height: int
    estimated_gpt56_original_tokens: int
    page_class: str
    extraction_action: str
    exclusion_reason: str | None


def _natural_key(path: Path) -> tuple[object, ...]:
    import re

    return tuple(int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", path.name))


def estimate_gpt56_original_tokens(width: int, height: int) -> int:
    """Estimate GPT-5.6 `original` image input tokens from its 32px patch count."""
    if width < 1 or height < 1:
        raise ValueError("image dimensions must be positive")
    return math.ceil(width / 32) * math.ceil(height / 32)


def inventory_sources(source_dir: Path, classes: dict[str, str] | None = None) -> list[SourceRecord]:
    """Inventory sources without copying, moving, or changing the originals."""
    try:
        from PIL import Image, ImageOps
    except ImportError as exc:  # pragma: no cover - dependency check
        raise RuntimeError("Pillow is required: pip install -e .[image]") from exc

    classes = classes or {}
    paths = sorted(
        (path for path in source_dir.iterdir() if path.is_file() and path.suffix.lower() in _IMAGE_SUFFIXES),
        key=_natural_key,
    )
    records: list[SourceRecord] = []
    for index, path in enumerate(paths, start=1):
        try:
            page_class = PageClass(classes.get(path.name, "unknown"))
        except ValueError as exc:
            raise ValueError(f"unsupported page class for {path.name}: {classes[path.name]}") from exc
        route = route_page(page_class)
        with Image.open(path) as image:
            image = ImageOps.exif_transpose(image)
            width, height = image.size
        records.append(
            SourceRecord(
                source_id=f"SRC-{index:04d}",
                source_file=path.name,
                source_path=str(path.resolve()),
                sha256=sha256_file(path),
                width=width,
                height=height,
                estimated_gpt56_original_tokens=estimate_gpt56_original_tokens(width, height),
                page_class=page_class.value,
                extraction_action=route.reason,
                exclusion_reason=route.reason if route.excluded_from_ocr else None,
            )
        )
    return records


def _write_xlsx(output_path: Path, records: Iterable[SourceRecord], project_name: str) -> None:
    records = list(records)
    book = Workbook()
    identity = book.active
    identity.title = "资料身份"
    identity.append(["项目", project_name])
    identity.append(["生成时间(UTC)", datetime.now(timezone.utc).isoformat()])
    identity.append(["发布状态", "blocked_pending_review"])

    inventory = book.create_sheet("来源页索引")
    inventory.append([
        "source_id", "source_file", "page_class", "action", "width", "height", "sha256",
        "gpt56_original_est_tokens", "source_path",
    ])
    for record in records:
        inventory.append([
            record.source_id, record.source_file, record.page_class, record.extraction_action,
            record.width, record.height, record.sha256, record.estimated_gpt56_original_tokens,
            record.source_path,
        ])

    excluded = book.create_sheet("排除页面")
    excluded.append(["source_id", "source_file", "page_class", "reason"])
    for record in records:
        if record.exclusion_reason:
            excluded.append([record.source_id, record.source_file, record.page_class, record.exclusion_reason])

    review = book.create_sheet("待审核")
    review.append(["source_id", "source_file", "reason", "blocking"])
    for record in records:
        if record.page_class == PageClass.UNKNOWN.value:
            review.append([record.source_id, record.source_file, "manual_classification_required", "yes"])

    for sheet in book.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(42, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    book.save(output_path)


def _write_reports(output_dir: Path, records: list[SourceRecord], project_name: str) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    classification_counts = {page_class.value: 0 for page_class in PageClass}
    for record in records:
        classification_counts[record.page_class] += 1
    unknown = classification_counts[PageClass.UNKNOWN.value]
    excluded = [record for record in records if record.exclusion_reason]
    total_tokens = sum(record.estimated_gpt56_original_tokens for record in records)
    release_status = "blocked_pending_review" if unknown else "not_released_missing_two_engine_consensus"

    (output_dir / "QC_REPORT.html").write_text(
        "<html><meta charset='utf-8'><body>"
        f"<h1>QC report: {project_name}</h1><p>release_status: {release_status}</p>"
        f"<p>source_count: {len(records)}; classified_count: {sum(classification_counts.values())}; "
        f"estimated_gpt56_original_tokens: {total_tokens}</p>"
        "<h2>Page classes</h2><ul>"
        + "".join(f"<li>{key}: {value}</li>" for key, value in classification_counts.items())
        + "</ul><h2>Excluded pages</h2><ul>"
        + "".join(f"<li>{record.source_id} {record.source_file}: {record.exclusion_reason}</li>" for record in excluded)
        + "</ul></body></html>",
        encoding="utf-8",
    )
    (output_dir / "completion_report.md").write_text(
        f"# {project_name} trial completion report\n\n"
        f"- Source images: {len(records)}\n"
        f"- Classified records: {sum(classification_counts.values())}\n"
        f"- Unknown pages: {unknown}\n"
        f"- Excluded pages: {len(excluded)}\n"
        f"- GPT-5.6 original-detail estimate: {total_tokens} image input tokens\n"
        f"- Release status: `{release_status}`\n\n"
        "This trial does not release numeric hydrology data: every unclassified page and every single-engine OCR candidate requires review.\n",
        encoding="utf-8",
    )
    with (output_dir / "token_budget.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["source_id", "source_file", "width", "height", "gpt56_original_est_tokens"])
        writer.writerows(
            (record.source_id, record.source_file, record.width, record.height, record.estimated_gpt56_original_tokens)
            for record in records
        )


def run_trial_audit(
    source_dir: Path,
    output_dir: Path,
    *,
    project_name: str,
    classes: dict[str, str] | None = None,
    master_workbook_path: Path | None = None,
) -> list[SourceRecord]:
    """Create the required auditable trial outputs from immutable source images."""
    records = inventory_sources(source_dir, classes)
    if not records:
        raise ValueError(f"no supported images found in {source_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "source_inventory.json").write_text(
        json.dumps([asdict(record) for record in records], ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _write_xlsx(output_dir / "QC_CHECKLIST.xlsx", records, project_name)
    if master_workbook_path:
        _write_xlsx(master_workbook_path, records, project_name)
    _write_reports(output_dir, records, project_name)
    return records


def validate_trial_release(output_dir: Path) -> ProjectValidation:
    """Validate the release gate for a completed trial without changing any files."""
    errors: list[str] = []
    warnings: list[str] = []
    required = ("source_inventory.json", "QC_REPORT.html", "QC_CHECKLIST.xlsx", "completion_report.md")
    for filename in required:
        if not (output_dir / filename).is_file():
            errors.append(f"missing required trial output: {filename}")
    inventory_path = output_dir / "source_inventory.json"
    if not inventory_path.is_file():
        return ProjectValidation(False, tuple(errors), tuple(warnings))
    try:
        records = json.loads(inventory_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        errors.append(f"invalid source_inventory.json: {exc}")
        return ProjectValidation(False, tuple(errors), tuple(warnings))
    if not records:
        errors.append("source inventory is empty")
    if any(not record.get("page_class") for record in records):
        errors.append("source count does not equal classified-page count")
    unknown_count = sum(record.get("page_class") == PageClass.UNKNOWN.value for record in records)
    if unknown_count:
        errors.append(f"unknown pages block release: {unknown_count}")
    return ProjectValidation(not errors, tuple(errors), tuple(warnings))
