from __future__ import annotations

import calendar
import json
import math
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from statistics import mean, median

from openpyxl import Workbook
from openpyxl.styles import Font

from .daily_candidate import OcrToken, read_ocr_tokens

_NUMERIC = re.compile(r"^\d+(?:\.\d+)?$")
_HYDROLOGICAL_STATES = (
    "部分河干", "部分渠干", "部分库干", "部分泊干", "部分洼干", "部分泽干",
    "部分封冻", "部分冰冻", "部分连底冻", "部分断流",
    "河干", "渠干", "库干", "泊干", "洼干", "泽干", "连底冻",
    "封冻", "冰冻", "断流", "无水", "停测",
)
_EXPLICIT_DITTO_MARKS = ('"', "“", "”", "〃", "″", "′", "•")


@dataclass(frozen=True)
class GridTemplate:
    """Geometry of a single 12-month matrix after perspective correction.

    Row positions deliberately include the printed blank separator after every
    five days.  This avoids treating an annual-yearbook matrix as a uniformly
    spaced 31-row grid.
    """

    month_centers: tuple[float, ...]
    first_row_y: tuple[float, ...]
    month_left: float
    month_right: float
    row_step: float
    five_day_gap: float
    cell_half_height: int = 17

    def __post_init__(self) -> None:
        if len(self.month_centers) != 12 or len(self.first_row_y) != 12:
            raise ValueError("a daily matrix needs 12 month centers and 12 first-row positions")


@dataclass(frozen=True)
class GridReading:
    month: int
    day: int
    value: float | None
    full_grid_text: str
    fixed_cell_text: str
    status: str


@dataclass(frozen=True)
class MonthlyCheck:
    month: int
    day_count: int
    calculated_mean: float | None
    calculated_maximum: float | None
    calculated_minimum: float | None
    printed_mean: float | None
    printed_maximum: float | None
    printed_minimum: float | None
    status: str


def row_y(template: GridTemplate, month_index: int, day: int) -> float:
    """Return the expected centre of a printed daily entry."""
    return (
        template.first_row_y[month_index]
        + template.row_step * (day - 1)
        + template.five_day_gap * ((day - 1) // 5)
    )


def _numeric(text: str) -> float | None:
    cleaned = text.strip().replace(" ", "")
    if not _NUMERIC.fullmatch(cleaned):
        return None
    return float(cleaned)


def _digits(text: str) -> str | None:
    """Return the significant digits only when OCR supplied a numeric token."""
    return text.strip().replace(" ", "").replace(".", "") if _numeric(text) is not None else None


def validate_month_local_rows(
    centers: tuple[float, ...], *, body_top: float, body_bottom: float
) -> bool:
    """Validate a complete 31-row physical grid for one month column."""
    if len(centers) != 31 or not body_top < centers[0] < centers[-1] < body_bottom:
        return False
    gaps = [right - left for left, right in zip(centers, centers[1:])]
    if any(gap <= 0 for gap in gaps):
        return False
    spacer_indexes = {4, 9, 14, 19, 24}
    ordinary = [gap for index, gap in enumerate(gaps) if index not in spacer_indexes]
    pitch = median(ordinary)
    if pitch <= 0 or not all(pitch * 0.55 <= gap <= pitch * 1.45 for gap in ordinary):
        return False
    spacers = [gaps[index] for index in sorted(spacer_indexes)]
    if not all(pitch * 1.15 <= gap <= pitch * 2.80 for gap in spacers):
        return False
    edge_gaps = (centers[0] - body_top, body_bottom - centers[-1])
    return all(pitch * 0.20 <= gap <= pitch * 2.50 for gap in edge_gaps)


def select_day_body_boundaries(
    strong_rules: tuple[float, ...],
    *,
    image_height: float,
    expected_header: float | None = None,
    expected_statistics: float | None = None,
) -> tuple[float, float] | None:
    """Select the physical header-bottom/statistics-top pair.

    The lower statistics run is the primary anchor.  This prevents a page
    title underline from becoming the day-body header when the photographed
    table is strongly tilted or curved.
    """
    height = float(image_height)
    rules = tuple(sorted(set(float(value) for value in strong_rules)))
    pairs = [
        (top, bottom)
        for top in rules
        for bottom in rules
        if height * 0.03 <= top <= height * 0.48
        and height * 0.48 <= bottom <= height * 0.94
        and height * 0.34 <= bottom - top <= height * 0.78
    ]
    if not pairs:
        return None

    statistics_anchor = (
        float(expected_statistics)
        if expected_statistics is not None
        and height * 0.48 <= float(expected_statistics) <= height * 0.94
        else None
    )

    def score(pair: tuple[float, float]) -> float:
        top, bottom = pair
        value = abs((bottom - top) / height - 0.57) * 4.0
        if statistics_anchor is not None:
            value += abs(bottom - statistics_anchor) / height * 10.0
        if (
            expected_header is not None
            and bottom - height * 0.78
            <= float(expected_header)
            <= bottom - height * 0.34
        ):
            value += abs(top - float(expected_header)) / height
        return value

    return min(pairs, key=score)


def state_month_statistics_closed(
    state: str,
    printed_statistics: tuple[object, ...],
    daily_values: tuple[object, ...],
    daily_raw: tuple[str, ...] | None = None,
) -> bool:
    """Prove a full hydrological-state month without reading tiny dittos.

    At least two independently printed monthly statistics must show the same
    state.  Existing daily values may be that state, unresolved, or a likely
    OCR reading of a tiny ditto as integer 0/1; decimals and other numeric
    observations block closure.
    """
    if state not in _HYDROLOGICAL_STATES:
        return False
    if sum(value == state for value in printed_statistics) < 2:
        return False
    raws = daily_raw or tuple("" for _ in daily_values)
    if len(raws) != len(daily_values):
        raise ValueError("daily_raw must align with daily_values")
    for value, raw in zip(daily_values, raws):
        if value is None or value == state:
            continue
        if isinstance(value, (int, float)) and float(value) in {0.0, 1.0} and "." not in raw:
            continue
        return False
    return True


def independent_model_majority(candidates: dict[str, object]) -> object | None:
    """Return a value only when two differently named models agree exactly."""
    groups: list[tuple[object, list[str]]] = []
    for model, value in candidates.items():
        if value is None:
            continue
        matched = next((group for group in groups if group[0] == value), None)
        if matched is None:
            groups.append((value, [model]))
        else:
            matched[1].append(model)
    winners = [group for group in groups if len(set(group[1])) >= 2]
    if len(winners) != 1:
        return None
    return winners[0][0]


def select_regular_footer_edges(
    strong_rules: tuple[float, ...], *, start: float, row_count: int
) -> tuple[float, ...] | None:
    """Select regular month-local footer rules while skipping isolated false lines."""
    rules = tuple(sorted(set(float(value) for value in strong_rules)))
    step_candidates = sorted({
        round(right - left, 3)
        for left, right in zip(rules, rules[1:])
        if 8 <= right - left <= 60
    })
    ranked: list[tuple[float, float, tuple[float, ...]]] = []
    for step in step_candidates:
        matched = []
        total_error = 0.0
        for index in range(row_count + 1):
            expected = start + index * step
            nearest = min(rules, key=lambda rule: abs(rule - expected))
            error = abs(nearest - expected)
            if error > max(3.0, step * 0.18):
                break
            matched.append(nearest)
            total_error += error
        if len(matched) == row_count + 1 and len(set(matched)) == row_count + 1:
            ranked.append((total_error, step, tuple(matched)))
    if not ranked:
        return None
    _, _, edges = min(ranked, key=lambda row: (row[0], row[1]))
    return edges


def is_single_vertical_date_glyph(
    components: tuple[tuple[int, int, int, int, int], ...],
    *,
    cell_width: int,
    cell_height: int,
) -> bool:
    """Accept only one tall interior component compatible with printed day 1."""
    usable = [
        (x, y, width, height, area)
        for x, y, width, height, area in components
        if area >= 3 and height >= cell_height * 0.25 and width < cell_width * 0.70
    ]
    if len(usable) != 1:
        return False
    _, _, width, height, area = usable[0]
    return (
        height >= cell_height * 0.42
        and height / max(width, 1) >= 1.45
        and width <= cell_width * 0.30
        and area <= cell_width * cell_height * 0.22
    )


def clamp_image_roi(
    bounds: tuple[int | float, int | float, int | float, int | float],
    image_shape: tuple[int, ...],
) -> tuple[int, int, int, int] | None:
    """Clamp a geometry-search region and reject a physically empty result."""
    height, width = image_shape[:2]
    x0, y0, x1, y1 = (round(value) for value in bounds)
    clamped = (
        max(0, min(width, x0)),
        max(0, min(height, y0)),
        max(0, min(width, x1)),
        max(0, min(height, y1)),
    )
    return clamped if clamped[0] < clamped[2] and clamped[1] < clamped[3] else None


def statistics_labels(variable: str) -> tuple[str, ...]:
    """Return only the footer rows printed by the daily-table family."""
    if variable in {"precipitation", "evaporation"}:
        return ("合计", "观测日数", "最大日量")
    if variable == "sediment_rate":
        return ("平均", "最大", "最大日期")
    return ("平均", "最大", "最大日期", "最小", "最小日期")


def strict_all_zero_statistics_proof(
    printed_means: tuple[float | None, ...],
    printed_maxima: tuple[float | None, ...],
) -> bool:
    """Require independent zero evidence in all 24 annual statistic cells."""
    return (
        len(printed_means) == 12
        and len(printed_maxima) == 12
        and all(value == 0 for value in printed_means)
        and all(value == 0 for value in printed_maxima)
    )


def protected_daily_source_state(status: str, value: object) -> bool:
    """Return whether month-local numeric rereading must skip this cell."""
    return (
        str(status).startswith("visual_")
        or status in {
            "source_blank", "source_slot_absent", "source_scan_missing",
            "source_blank_visual_verified",
        }
        or isinstance(value, str)
    )


def source_state_evidence(
    raw: str, *, explicit_ditto: bool = False, source_blank: bool = False
) -> str | None:
    """Return the only source evidence that may support a state-valued cell.

    A printed dash, underline or empty crop terminates state propagation. It
    never behaves like a ditto mark, even when the preceding day is a state.
    """
    if source_blank:
        return None
    compact = re.sub(r"\s+", "", str(raw or ""))
    if any(state in compact for state in _HYDROLOGICAL_STATES):
        return "direct_state"
    if explicit_ditto or any(mark in compact for mark in _EXPLICIT_DITTO_MARKS):
        return "explicit_ditto"
    return None


def statistics_source_conflict(
    *, existing_raw: str, existing_value: object,
    candidate_raw: str, candidate_value: object,
    label: str | None = None, year: int | None = None,
    month: int | None = None, candidate_score: float | None = None,
    date_score_threshold: float = 0.98,
) -> str | None:
    """Return the source-preservation reason that blocks a reread candidate."""
    if (
        label not in {"最大日期", "最小日期"}
        and isinstance(existing_value, str)
        and existing_value.strip()
        and _numeric(existing_raw) is None
    ):
        return "existing_hydrological_state"
    if _numeric(existing_raw) is not None and _numeric(candidate_raw) is not None:
        old_digits = re.sub(r"\D", "", existing_raw)
        new_digits = re.sub(r"\D", "", candidate_raw)
        if (
            existing_value != candidate_value
            and old_digits == new_digits
            and "." in existing_raw.replace(",", ".")
            and "." not in candidate_raw.replace(",", ".")
        ):
            return "candidate_drops_existing_decimal"
    if (
        label in {"最大日期", "最小日期"}
        and year is not None
        and month is not None
        and candidate_score is not None
        and candidate_score < date_score_threshold
    ):
        old = _numeric(str(existing_value))
        new = _numeric(str(candidate_value))
        last_day = calendar.monthrange(year, month)[1]
        if (
            old is not None and new is not None and old != new
            and old.is_integer() and new.is_integer()
            and 1 <= int(old) <= last_day and 1 <= int(new) <= last_day
        ):
            return "low_confidence_plausible_date_conflict"
    return None


def expand_compact_water_level(raw: str, prefix: int | None) -> tuple[float | None, int | None]:
    """Expand one printed water-level token without consulting the monthly mean."""
    match = re.search(r"[+\-]?\d+(?:[\.,]\d+)?", str(raw))
    if not match:
        return None, prefix
    token = match.group(0).replace(",", ".")
    if "." in token:
        value = float(token)
        return value, math.floor(value)
    digits = re.sub(r"\D", "", token)
    if len(digits) <= 2:
        if prefix is None:
            return None, None
        return round(prefix + int(digits) / 100.0, 6), prefix
    if len(digits) in {3, 4}:
        value = int(digits) / 100.0
        return value, math.floor(value)
    return None, prefix


def _fit_row_centers(
    tokens: list[OcrToken], template: GridTemplate, year: int
) -> tuple[tuple[float, ...], ...]:
    """Fit a small per-column correction from full-grid OCR coordinates.

    Full-grid OCR is used only for geometry registration here.  Its text is
    independently retained for later agreement checks; it is never silently
    treated as the fixed-cell result.
    """
    try:
        import numpy as np
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("numpy is required for grid registration") from exc

    output: list[tuple[float, ...]] = []
    for column in range(12):
        valid_days = range(1, calendar.monthrange(year, column + 1)[1] + 1)
        expected = {day: row_y(template, column, day) for day in valid_days}
        candidates = [
            token
            for token in tokens
            if _numeric(token.text) is not None
            and abs(token.x - template.month_centers[column]) <= 105
        ]
        assigned: dict[int, OcrToken] = {}
        for token in candidates:
            day = min(expected, key=lambda candidate: abs(token.y - expected[candidate]))
            if abs(token.y - expected[day]) > 24:
                continue
            previous = assigned.get(day)
            if previous is None or abs(token.y - expected[day]) < abs(previous.y - expected[day]):
                assigned[day] = token

        # A line plus a five-day-gap term captures remaining perspective drift.
        if len(assigned) >= 8:
            matrix = np.array(
                [[1.0, day - 1, (day - 1) // 5] for day in assigned], dtype=float
            )
            values = np.array([assigned[day].y for day in assigned], dtype=float)
            coefficients, *_ = np.linalg.lstsq(matrix, values, rcond=None)
            residuals = values - matrix @ coefficients
            keep = abs(residuals) <= 12
            if int(keep.sum()) >= 6:
                coefficients, *_ = np.linalg.lstsq(matrix[keep], values[keep], rcond=None)
            row_positions = tuple(
                float(coefficients @ np.array([1.0, day - 1, (day - 1) // 5]))
                for day in range(1, 32)
            )
        else:
            row_positions = tuple(row_y(template, column, day) for day in range(1, 32))
        output.append(row_positions)
    return tuple(output)


def _map_full_grid_text(
    tokens: list[OcrToken],
    template: GridTemplate,
    year: int,
    row_centers: tuple[tuple[float, ...], ...],
) -> dict[tuple[int, int], str]:
    mapped: dict[tuple[int, int], tuple[OcrToken, float]] = {}
    for token in tokens:
        if _numeric(token.text) is None:
            continue
        column = min(range(12), key=lambda index: abs(token.x - template.month_centers[index]))
        if abs(token.x - template.month_centers[column]) > 105:
            continue
        max_day = calendar.monthrange(year, column + 1)[1]
        day = min(range(1, max_day + 1), key=lambda value: abs(token.y - row_centers[column][value - 1]))
        distance = abs(token.y - row_centers[column][day - 1])
        if distance > 17:
            continue
        key = (column + 1, day)
        previous = mapped.get(key)
        if previous is None or distance < previous[1]:
            mapped[key] = (token, distance)
    return {key: token.text for key, (token, _) in mapped.items()}


def _cell_bounds(template: GridTemplate, column: int, center_y: float) -> tuple[int, int, int, int]:
    centers = template.month_centers
    if column == 0:
        left = template.month_left
    else:
        left = (centers[column - 1] + centers[column]) / 2
    if column == 11:
        right = template.month_right
    else:
        right = (centers[column] + centers[column + 1]) / 2
    return (
        round(left + 7),
        round(center_y - template.cell_half_height),
        round(right - 7),
        round(center_y + template.cell_half_height),
    )


def extract_grid_daily(
    rectified_image: Path, *, year: int, template: GridTemplate
) -> tuple[list[GridReading], dict[str, object]]:
    """Read a perspective-corrected daily matrix one printed cell at a time."""
    try:
        import cv2
        import numpy as np
        from rapidocr_onnxruntime import RapidOCR
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("OpenCV, numpy and RapidOCR are required") from exc

    encoded = np.fromfile(rectified_image, dtype=np.uint8)
    image = cv2.imdecode(encoded, cv2.IMREAD_COLOR)
    if image is None:
        raise ValueError(f"cannot decode image: {rectified_image}")

    full_tokens = read_ocr_tokens(rectified_image)
    row_centers = _fit_row_centers(full_tokens, template, year)
    full_grid = _map_full_grid_text(full_tokens, template, year, row_centers)
    engine = RapidOCR()
    readings: list[GridReading] = []

    for column in range(12):
        for day in range(1, calendar.monthrange(year, column + 1)[1] + 1):
            x0, y0, x1, y1 = _cell_bounds(template, column, row_centers[column][day - 1])
            if x0 < 0 or y0 < 0 or x1 > image.shape[1] or y1 > image.shape[0]:
                raise ValueError("registered cell bounds fall outside the rectified image")
            cell = image[y0:y1, x0:x1]
            cell = cv2.resize(cell, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC)
            cell = cv2.copyMakeBorder(
                cell, 24, 24, 24, 24, cv2.BORDER_CONSTANT, value=(255, 255, 255)
            )
            result, _ = engine(cell, use_det=False)
            fixed_text = str(result[0][0]).strip() if result else ""
            # A slightly taller retry recovers entries whose baseline drifts by
            # a few pixels near a curved photographed page.  It is only used
            # when the tight cell crop is not numeric.
            if _numeric(fixed_text) is None:
                x0, y0, x1, y1 = _cell_bounds(template, column, row_centers[column][day - 1])
                wider_bounds = clamp_image_roi((x0, y0 - 6, x1, y1 + 6), image.shape)
                if wider_bounds is None:
                    raise ValueError("expanded cell bounds are physically empty")
                wx0, wy0, wx1, wy1 = wider_bounds
                wider = image[wy0:wy1, wx0:wx1]
                wider = cv2.resize(wider, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC)
                wider = cv2.copyMakeBorder(
                    wider, 24, 24, 24, 24, cv2.BORDER_CONSTANT, value=(255, 255, 255)
                )
                retry, _ = engine(wider, use_det=False)
                retry_text = str(retry[0][0]).strip() if retry else ""
                if _numeric(retry_text) is not None:
                    fixed_text = retry_text
            full_text = full_grid.get((column + 1, day), "")
            fixed_value = _numeric(fixed_text)
            full_value = _numeric(full_text)
            if fixed_value is not None and full_value is not None and math.isclose(fixed_value, full_value):
                value, status = fixed_value, "dual_read_agree"
            elif _digits(fixed_text) is not None and _digits(fixed_text) == _digits(full_text):
                # The two readings agree on every digit; prefer the version
                # retaining an explicit decimal point for a correctly typed
                # spreadsheet value.
                value = fixed_value if "." in fixed_text else full_value
                status = "dual_read_format_normalized"
            elif (
                fixed_value is not None
                and full_value is not None
                and _digits(fixed_text).endswith(_digits(full_text))
            ):
                # A fixed-cell crop preserves the full-grid suffix but adds a
                # leading digit (for example 2.27 versus a truncated 27).
                value, status = fixed_value, "cell_recovers_truncated_full_grid"
            elif fixed_value is not None:
                value, status = fixed_value, "cell_only_pending_statistical_check"
            elif full_value is not None:
                value, status = full_value, "full_grid_only_pending_cell_review"
            else:
                value, status = None, "unreadable"
            readings.append(GridReading(column + 1, day, value, full_text, fixed_text, status))

    registration = {
        "rectified_image": str(rectified_image),
        "month_centers": list(template.month_centers),
        "fitted_row_centers": [list(column) for column in row_centers],
        "full_grid_token_count": len(full_tokens),
    }
    return readings, registration


def monthly_checks(
    readings: list[GridReading],
    *,
    year: int,
    printed_means: tuple[float, ...],
    printed_maxima: tuple[float, ...],
    printed_minima: tuple[float, ...],
    require_extrema: bool = True,
    mean_rule: str = "arithmetic",
    extrema_rule: str = "instantaneous_bounds",
) -> list[MonthlyCheck]:
    if mean_rule not in {"arithmetic", "observed"}:
        raise ValueError("mean_rule must be 'arithmetic' or 'observed'")
    if extrema_rule not in {"instantaneous_bounds", "daily_mean_extrema"}:
        raise ValueError("extrema_rule must be 'instantaneous_bounds' or 'daily_mean_extrema'")
    if not all(len(series) == 12 for series in (printed_means, printed_maxima, printed_minima)):
        raise ValueError("all printed statistics need 12 monthly values")
    result: list[MonthlyCheck] = []
    for month in range(1, 13):
        values = [reading.value for reading in readings if reading.month == month]
        expected_count = calendar.monthrange(year, month)[1]
        complete = len(values) == expected_count and all(value is not None for value in values)
        numeric = [value for value in values if value is not None]
        calculated_mean = mean(numeric) if complete else None
        calculated_maximum = max(numeric) if complete else None
        calculated_minimum = min(numeric) if complete else None
        printed_mean = printed_means[month - 1]
        printed_maximum = printed_maxima[month - 1]
        printed_minimum = printed_minima[month - 1]
        def printed_uncertainty(value: float) -> float:
            magnitude = abs(float(value))
            if magnitude == 0:
                return 0.0
            places = 2 - math.floor(math.log10(magnitude))
            return 0.5 * 10 ** (-places)

        # Compare overlapping three-significant-digit print intervals. The
        # statistic may be calculated from observations before daily values
        # are rounded for publication.
        tolerance = printed_uncertainty(printed_mean) + (
            sum(printed_uncertainty(value) for value in numeric) / expected_count
            if complete else 0.0
        ) + 1e-12
        mean_passed = complete and (
            mean_rule == "observed" or abs(calculated_mean - printed_mean) <= tolerance
        )
        if extrema_rule == "daily_mean_extrema":
            extrema_passed = complete and (
                math.isclose(calculated_maximum, printed_maximum, abs_tol=1e-9)
                and math.isclose(calculated_minimum, printed_minimum, abs_tol=1e-9)
            )
        else:
            maximum_tolerance = printed_uncertainty(printed_maximum) + max(
                (printed_uncertainty(value) for value in numeric), default=0.0
            )
            minimum_tolerance = printed_uncertainty(printed_minimum) + max(
                (printed_uncertainty(value) for value in numeric), default=0.0
            )
            extrema_passed = complete and (
                calculated_maximum <= printed_maximum + maximum_tolerance
                and calculated_minimum >= printed_minimum - minimum_tolerance
            )
        passed = mean_passed and (
            extrema_passed or not require_extrema
        )
        result.append(
            MonthlyCheck(
                month,
                expected_count,
                calculated_mean,
                calculated_maximum,
                calculated_minimum,
                printed_mean,
                printed_maximum,
                printed_minimum,
                "passed" if passed else "needs_visual_review",
            )
        )
    return result


def write_grid_audit(path: Path, *, readings: list[GridReading], registration: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            {"registration": registration, "readings": [asdict(reading) for reading in readings]},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def write_clean_daily_workbook(
    path: Path,
    *,
    basin: str,
    year: int,
    station: str,
    source_file: str,
    readings: list[GridReading],
    checks: list[MonthlyCheck],
) -> None:
    """Write the deliberately compact delivery workbook, separate from audit JSON."""
    if any(check.status != "passed" for check in checks):
        raise ValueError("monthly validation has not passed; do not write a clean delivery workbook")
    if any(reading.value is None for reading in readings):
        raise ValueError("unreadable cells cannot enter a clean delivery workbook")

    book = Workbook()
    daily = book.active
    daily.title = "逐日数据"
    daily.append(["日"] + [f"{month}月" for month in range(1, 13)])
    lookup = {(reading.day, reading.month): reading.value for reading in readings}
    for day in range(1, 32):
        daily.append(
            [day]
            + [lookup.get((day, month)) if day <= calendar.monthrange(year, month)[1] else None for month in range(1, 13)]
        )

    checks_sheet = book.create_sheet("月复核")
    checks_sheet.append(["月", "天数", "计算均值", "印刷均值", "计算最大", "计算最小", "结果"])
    for check in checks:
        checks_sheet.append(
            [
                check.month,
                check.day_count,
                check.calculated_mean,
                check.printed_mean,
                check.calculated_maximum,
                check.calculated_minimum,
                "通过",
            ]
        )

    metadata = book.create_sheet("说明")
    metadata.append(["字段", "内容"])
    metadata.append(["流域", basin])
    metadata.append(["年份", year])
    metadata.append(["站点", station])
    metadata.append(["来源图片", source_file])
    metadata.append(["复核", "逐格识别、整表定位读数、印刷月统计三重复核均通过"])
    metadata.append(["无效日期", "空白；按该年份实际月天数处理"])

    for sheet in book.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = max(11, min(30, max(len(str(cell.value or "")) for cell in column) + 2))
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.###"
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
